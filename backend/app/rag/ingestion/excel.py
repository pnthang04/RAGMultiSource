import pandas as pd
import os
import json
import re
from io import BytesIO
from openpyxl import load_workbook, Workbook
from markitdown import MarkItDown
import xlrd

def _read_base_excel(path: str):
    df = pd.read_excel(path)
    descriptions = []
    for _, row in df.iterrows():
        desc = " . ".join([f"{col} là {row[col]}" for col in df.columns])
        descriptions.append(desc)
    return "\n".join(descriptions)

def convert_xls_to_xlsx(xls_path):
    xls_wb = xlrd.open_workbook(xls_path)
    new_wb = Workbook()
    new_wb.remove(new_wb.active)   

    for sheet_idx in range(xls_wb.nsheets):
        xls_ws = xls_wb.sheet_by_index(sheet_idx)
        new_ws = new_wb.create_sheet(title=xls_ws.name)

        for row_idx in range(xls_ws.nrows):
            row_data = []
            for col_idx in range(xls_ws.ncols):
                cell = xls_ws.cell(row_idx, col_idx)
                row_data.append(cell.value if cell.ctype != 0 else None)
            new_ws.append(row_data)

    xlsx_path = xls_path.replace(".xls", "_converted.xlsx")
    new_wb.save(xlsx_path)
    return xlsx_path

def is_empty(sequence):
    return all(cell is None or str(cell).strip() == "" for cell in sequence)

def _split_table_recursive(table):
    row_blocks = []
    current_block = []
    for row in table:
        if is_empty(row):
            if current_block:
                row_blocks.append(current_block)
                current_block = []
        else:
            current_block.append(row)
    if current_block:
        row_blocks.append(current_block)
        
    if len(row_blocks) > 1:
        result = []
        for block in row_blocks:
            result.extend(_split_table_recursive(block))
        return result
        
    transposed = list(zip(*table))
    col_blocks = []
    current_col_block = []
    for col in transposed:
        if is_empty(col):
            if current_col_block:
                col_blocks.append(current_col_block)
                current_col_block = []
        else:
            current_col_block.append(col)
    if current_col_block:
        col_blocks.append(current_col_block)
        
    if len(col_blocks) > 1:
        result = []
        for col_block in col_blocks:
            block = list(zip(*col_block))
            result.extend(_split_table_recursive(block))
        return result
        
    return [table]

def extract_tables_from_sheet(ws):
    table = list(ws.iter_rows(values_only=True))
    if not table:
        return []
    return _split_table_recursive(table)


def table_to_markdown(table, md):
    temp_wb = Workbook()
    temp_ws = temp_wb.active
    temp_ws.append([None] * len(table[0]))

    for row in table:
        temp_ws.append(list(row))

    buffer = BytesIO()
    temp_wb.save(buffer)
    buffer.seek(0)

    result = md.convert(buffer)
    return result.text_content


def _read_to_markdown(file_path):
    file_name = os.path.basename(file_path)

    if file_path.lower().endswith(".xls"):
        file_path = convert_xls_to_xlsx(file_path)

    wb = load_workbook(file_path, data_only=True)
    md = MarkItDown()

    metadata_list = []
    markdown_sections = []

    for ws in wb.worksheets:
        sheet_md = [f"# Sheet: {ws.title}\n"]
        tables = extract_tables_from_sheet(ws)

        for table in tables:
            row_count = len(table)
            col_count = len(table[0]) if row_count else 0

            metadata_list.append({
                "sheet_name": ws.title,
                "rows": row_count,
                "columns": col_count,
                "source_file": file_name,
            })

            table_md = table_to_markdown(table, md)
            sheet_md.append(f"{table_md}\n\n---\n")

        markdown_sections.append("\n".join(sheet_md))

    metadata_str = json.dumps(metadata_list, ensure_ascii=False, indent=2)

    final_md_content = (
        "## Metadata\n\n"
        f"{metadata_str}\n\n"
        "---\n\n"
        + "\n".join(markdown_sections)
    )

    pattern = r'\|.*Unnamed:.*\|\n\|[ \-\|]*\|\n?'
    final_md_content = re.sub(pattern, '', final_md_content)

    return final_md_content

def read_full_excel(path: str, to_markdown=True):
    if to_markdown:
        return _read_to_markdown(path)
    
    return _read_base_excel(path)

def read_excel(path: str, k=5):
    try:
        df = pd.read_excel(path)
    except UnicodeDecodeError:
        df = pd.read_excel(path, encoding="latin1")
    
    result = {
        "metadata": {
            "file_name": os.path.basename(path),
            "file_size": os.path.getsize(path)
        },
        "num_columns": len(df.columns),
        "num_rows": len(df),
        "first_k_rows": df.head(k).to_dict(orient="records")
    }
    return json.dumps(result, ensure_ascii=False, indent=2)